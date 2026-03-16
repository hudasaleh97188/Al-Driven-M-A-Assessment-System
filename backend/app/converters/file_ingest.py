"""
file_ingest.py
--------------
Multi-format file ingestion for the DealLens extraction pipeline.

Supported formats and handling strategy:
  - PDF   → passed directly to Gemini as application/pdf
  - PPTX  → passed directly to Gemini as application/vnd.openxmlformats-officedocument.presentationml.presentation
  - CSV   → passed directly to Gemini as text/csv
  - XLSX  → pre-processed with openpyxl into structured text, then passed as text/plain

Each public function returns a list of ``ContentPart`` named-tuples that the
extractor can feed straight into the Gemini ``contents`` list.
"""

from __future__ import annotations

import csv
import io
import os
from dataclasses import dataclass
from typing import List

from loguru import logger

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover – optional at import time
    load_workbook = None  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# Public data structure
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class ContentPart:
    """Carrier for a single piece of content destined for the Gemini API."""

    data: bytes
    mime_type: str
    source_filename: str


# ---------------------------------------------------------------------------
# MIME-type mapping
# ---------------------------------------------------------------------------

_EXT_TO_MIME = {
    ".pdf":  "application/pdf",
    ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ".csv":  "text/csv",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

SUPPORTED_EXTENSIONS = set(_EXT_TO_MIME.keys())


# ---------------------------------------------------------------------------
# Excel → structured text
# ---------------------------------------------------------------------------

def _xlsx_to_structured_text(file_path: str) -> str:
    """
    Read an Excel workbook with *openpyxl* and convert every sheet into a
    human-readable text block that preserves headers and row relationships.

    Returns a single UTF-8 string suitable for ``text/plain`` ingestion.
    """
    if load_workbook is None:
        raise ImportError(
            "openpyxl is required for Excel ingestion. "
            "Install it with: pip install openpyxl"
        )

    wb = load_workbook(file_path, data_only=True, read_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # First non-empty row is treated as the header
        header_idx = 0
        for i, row in enumerate(rows):
            if any(cell is not None for cell in row):
                header_idx = i
                break

        headers = [str(c) if c is not None else "" for c in rows[header_idx]]

        lines: list[str] = [f"=== Sheet: {sheet_name} ==="]
        lines.append(" | ".join(headers))
        lines.append("-" * (len(" | ".join(headers))))

        for row in rows[header_idx + 1:]:
            if all(cell is None for cell in row):
                continue
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))

        sections.append("\n".join(lines))

    wb.close()
    return "\n\n".join(sections)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def ingest_file(file_path: str) -> ContentPart:
    """
    Ingest a single file and return a ``ContentPart`` ready for the Gemini API.

    Raises ``ValueError`` for unsupported file extensions.
    """
    ext = os.path.splitext(file_path)[1].lower()
    filename = os.path.basename(file_path)

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type '{ext}' for {filename}. "
            f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    if ext in (".xlsx", ".xls"):
        # Excel → structured text
        logger.info(
            "[INGEST] Converting Excel file to structured text: {}", filename
        )
        text = _xlsx_to_structured_text(file_path)
        return ContentPart(
            data=text.encode("utf-8"),
            mime_type="text/plain",
            source_filename=filename,
        )

    # PDF, PPTX, CSV → read raw bytes and pass with native MIME type
    mime = _EXT_TO_MIME[ext]
    with open(file_path, "rb") as fh:
        raw = fh.read()

    size_mb = len(raw) / 1_048_576
    logger.info(
        "[INGEST] Loaded {} ({}) – {:.1f} MB", filename, mime, size_mb
    )
    return ContentPart(data=raw, mime_type=mime, source_filename=filename)


def ingest_files(file_paths: List[str]) -> List[ContentPart]:
    """
    Convenience wrapper: ingest multiple files, skipping any that fail,
    and return the list of successfully ingested ``ContentPart`` objects.
    """
    parts: list[ContentPart] = []
    for fp in file_paths:
        if not os.path.exists(fp):
            logger.warning("[INGEST] File not found, skipping: {}", fp)
            continue
        try:
            parts.append(ingest_file(fp))
        except Exception as exc:
            logger.error("[INGEST] Failed to ingest {}: {}", fp, exc)
    return parts
